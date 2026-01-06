from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
import logging

logger = logging.getLogger(__name__)

class ExcelTemplateProcessor:
    """
    Excelテンプレート処理
    {{氏名}} などのプレースホルダーを置換します。
    """
    
    def __init__(self, template_path: str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found: {template_path}")
        self.wb = load_workbook(template_path)
        self.ws = self.wb.active
    
    def fill_placeholders(self, data: dict):
        """
        プレースホルダーを置換
        例: {{氏名}} → NGUYEN VAN A
        """
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    new_value = original_value
                    for key, value in data.items():
                        placeholder = f"{{{{{key}}}}}"
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, str(value or ""))
                    
                    if new_value != original_value:
                        cell.value = new_value
    
    def fill_cell(self, cell_ref: str, value: any):
        """特定セルに値を設定 (A1など)"""
        self.ws[cell_ref] = value
    
    def save(self, output_path: str):
        """保存"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        return output_path
