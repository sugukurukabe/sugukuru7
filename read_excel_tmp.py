
import openpyxl
import sys

def read_excel(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"File: {path}")
        print(f"Sheets: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            # Print first 20 rows
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 40: break
                row_str = " | ".join([str(c) if c is not None else "" for c in row])
                print(f"{i+1}: {row_str}")
    except Exception as e:
        print(f"Error reading {path}: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_excel(sys.argv[1])
